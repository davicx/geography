import openpyxl
import openpyxl
from pathlib import Path


def get_hyperlink():
    cwd = Path.cwd()
    excel_file = cwd / 'examples' / 'excelHandling' /  'excel' / 'hyperink.xlsx'

    #Open with full path
    #excelBook = openpyxl.load_workbook('/Users/david/Desktop/David/www/geography/examples/excelHandling/excel/hyperink.xlsx')

    #Open with pathlib
    excelBook = openpyxl.load_workbook(excel_file)

    excelSheet = excelBook['News']
    
    return excelSheet.cell(row=2, column=1).hyperlink.target

hyperlink = get_hyperlink()
print("Hyperlink")
print(hyperlink)
print(" ")
